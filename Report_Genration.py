from flask import Flask, render_template, flash, url_for, request, make_response, jsonify, session, send_from_directory
from werkzeug.utils import secure_filename
import os, time
import io
import base64
import json
import datetime
import xlsxwriter
import os, sys, glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles import Border, Side
# from flask import send_from_directory
from flask import send_file
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from fnmatch import fnmatch
import time
from db_dervices import DBO
import pandas as pd
import numpy as np
import datetime
import random
from dateutil.tz import gettz
import math
from openpyxl.styles import Alignment

dbo          =  DBO()
MYDIR        = os.path.dirname(__file__)
REGULAR_SIZE = 11
REGULAR_FONT = 'Cambria'
server       = 'smtp.gmail.com'
port         = 587
username     = "aajeetshk@gmail.com"
password     = "ilbumnmnsnqletdk"
send_from    = "aajeetshk@gmail.com"
send_to      = "ashish@pinpointengineers.co.in"


def sum_velocty(row):
    sum = float(row.V1) + float(row.V2) + float(row.V3) + float(row.V4) + float(row.V5)
    avg = sum / 5
    return float(avg)


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            



class Report_Genration:
    def __init__(self):
        self = self

    @staticmethod
    def generate_report_air_velocity(data, basic_details,user,userid):
        sr_no         = basic_details['sr_no']
        company_name  = basic_details['company_name']
        room_volume   = basic_details['room_volume']
        room_name     = basic_details['room_name']
        ahu_number    = basic_details['ahu_number']
        test_taken    = datetime.datetime.today().strftime('%d/%m/%Y')
        test_taken    = basic_details['Test_taken']
        location      = basic_details['Department']
        grade         = basic_details['grade']
        acph_thresold = basic_details['acph_thresold']
        company_name_val = company_name
        SR_NO_val     = sr_no
        test_date     = test_taken

        temp_df             = dbo.get_company_details_by_company_name(company_name)
        report_number       = temp_df.REPORT_NUMBER.values[0]
        customer_address    = temp_df.ADDRESS.values[0]
        
        temp_df             = dbo.get_equipment(SR_NO_val) 
        INSTRUMENT_NAME     = temp_df.EQUIPMENT_NAME.values[0]
        MAKE                = temp_df.MAKE.values[0]
        MAKE_MODEL          = temp_df.MODEL_NUMBER.values[0]
        done_date           = temp_df.DONE_DATE.values[0].split()[0]
        due_date            = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY            = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        
        Nature_of_test      = "AIR VELOCITY"
        customer_name       = company_name_val
        Test_taken          = test_date
        SRNO                = SR_NO_val

        compan_name         = company_name_val
        compan_name         = compan_name.replace(".", "")
        compan_name         = compan_name.replace("/", "")
        compan_name         = compan_name.replace(" ", "")
        
        report_counter  = dbo.get_report_number(Nature_of_test,Test_taken,str(company_name_val))
        report_counter  = report_counter+1
        temp_date       = Test_taken.replace("/","")        
        report_number   = str("PPE{}AV{}{}".format(report_number,temp_date,report_counter))
        
        working_directory   = MYDIR + "/" +"static/Report/AIR_VELOCITY_REPORT/{}"
        final_working_directory = "static/Report/AIR_VELOCITY_REPORT/{}/{}.xlsx"
        file_name           = str(report_number)+"xlsx" #"{}_AIR_VELOCOTY_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));

        store_location          = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR + "/" +final_working_directory.format(compan_name, file_name)
        print(final_working_directory)

 
        wb = load_workbook(os.path.join("static/inputData/Template/",'Air_velocity_template.xlsx'))
        ws = wb.active
        ws.protection.sheet = True

        if grade=="A" or grade=="ISO 5":
            wb = load_workbook(os.path.join("static/inputData/Template/",'Air_velocity_template_grade_a.xlsx'))
            ws = wb.active
            ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        #ws['F5'] = str(Nature_of_test)
        ws['F5'] = str(Test_taken)
        ws['F6'] = str(ahu_number)
        ws['F7'] = str(location)
        

        
        ws['F8']  = report_number
        ws['B14'] = str(INSTRUMENT_NAME)
        ws['E14'] = str(MAKE_MODEL)
        ws['I14'] = str(SRNO)
        ws['M14'] = str(VALIDITY.replace("-", "/"))

        ws.merge_cells(start_row=19, start_column=2, end_row=19, end_column=3)
        ws.merge_cells(start_row=19, start_column=15, end_row=19, end_column=16)

        row = 20
        data['AVG_Velocity'] = data.apply(sum_velocty, axis=1)
        if (grade=="A") or (grade=="ISO 5"):
            data['CFM'] = data.apply(lambda x: int(x['AVG_Velocity']), axis=1)
            Total_cfm   = data['CFM'].sum()
            Total_cfm   = round(Total_cfm,2)
            ACPH_VALUE  = float((Total_cfm * 60))
            ACPH_VALUE  = int(ACPH_VALUE)
        else:
            data['CFM'] = data.apply(lambda x: float(x['AVG_Velocity']) * float(x['Inlet_size']), axis=1)
            data['CFM'] = data['CFM'].round(2)
            Total_cfm   = data['CFM'].sum()
            Total_cfm   = round(Total_cfm,2)
            ACPH_VALUE  = float((Total_cfm * 60)) / float(room_volume)
            ACPH_VALUE  = round(ACPH_VALUE,2)

        for row_data in data.itertuples():
            
          
            if (grade=="A") or (grade=="ISO 5"):
                ws["D" + str(row)] = row_data.Label_number
                ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
                ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
                ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)
                ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)
                ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=14)
                ws.merge_cells(start_row=row, start_column=15, end_row=row, end_column=16)
                
                ws["E" + str(row)] = row_data.V1
                ws["G" + str(row)] = row_data.V2
                ws["I" + str(row)] = row_data.V3
                ws["K" + str(row)] = row_data.V4
                ws["M" + str(row)] = row_data.V5
                ws["O" + str(row)] = row_data.AVG_Velocity
            else:
                ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
                ws["D" + str(row)]    = row_data.Label_number
                currentCell           = ws["D" + str(row)]
                currentCell.alignment = Alignment(horizontal='center', vertical='center')
                ws["F" + str(row)] = row_data.V1
                ws["G" + str(row)] = row_data.V2
                ws["H" + str(row)] = row_data.V3
                ws["I" + str(row)] = row_data.V4
                ws["J" + str(row)] = row_data.V5
                ws["K" + str(row)] = row_data.AVG_Velocity
                ws["L" + str(row)] = row_data.Inlet_size
                ws["M" + str(row)] = row_data.CFM
        
            row += 1

        ws.merge_cells(start_row=20, start_column=2, end_row=row - 1, end_column=3)
        
        ws['B20'] = room_name
        
        if (grade=="A") or (grade=="ISO 5"):
            pass
        else:
            ws.merge_cells(start_row=20, start_column=14, end_row=row - 1, end_column=14)
            ws['N20'] = room_volume
            ws.merge_cells(start_row=20, start_column=15, end_row=row - 1, end_column=16)
            ws['O20'] = ACPH_VALUE

        currentCell = ws['B20']
        currentCell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

        currentCell = ws['N20']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        currentCell = ws['O20']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        if (grade=="A") or (grade=="ISO 5"):
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        if (grade!="A") and (grade!="ISO 5"):
           ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
           ws["B" + str(row)] = "TOTAL CFM"
           currentCell = ws["B" + str(row)]
           currentCell.alignment = Alignment(horizontal='right', vertical='center')

           ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=13)
           ws["M" + str(row)] = Total_cfm
           ws.merge_cells(start_row=row, start_column=15, end_row=row, end_column=16)

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        if (grade=="A") or (grade=="ISO 5"):
            ws["B" + str(row)] = "Acceptance criteria : 90\u00B1 20 % ( 72 TO 108 FPM)".format(str(acph_thresold))
        else: 
            ws["B" + str(row)] = "Acceptance criteria :Not less Than {} ACPH".format(str(acph_thresold))

        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=16)
        if 'ISO' in grade:
            ws["J" + str(row)] = "CLASS : {}".format(grade)
        else:
            ws["J" + str(row)] = "GRADE : {}".format(grade)

        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)

        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 2)] = "REPORT PREPARED BY"

        currentCell = ws["B" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 2)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)
        
        testtime = datetime.datetime.now(tz=gettz('Asia/Kolkata'))
        testtime = str(testtime.strftime('%d/%m/%Y %H:%M:%S')).split(" ")[0]
        ws["B" + str(row + 3)]= "Electronically Signed by {} on {} ".format(userid, testtime)
        currentCell = ws["B" + str(row + 3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEERS"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)
        
       
        testtime = datetime.datetime.now(tz=gettz('Asia/Kolkata'))
        testtime = str(testtime.strftime('%d/%m/%Y %H:%M:%S')).split(".")[0]

        #ws["B" + str(row + 9)] = "Report Prepared by {} on {}".format(user, testtime)
        #ws["B" + str(row + 9)] = "Report Prepared by {} ".format(user)
        set_border(ws, 'B1:P' + str(row + 8))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)
        dbo.insert_file(Nature_of_test,report_number,userid,Test_taken,testtime,str(company_name_val),location,final_working_directory)
        return file_name, store_location
        
        
    @staticmethod
    def generate_report_pao(data, basic_details,user,userid):
    
        test_type        = basic_details['test_type']
        sr_no            = basic_details['sr_no']
        company_name     = basic_details['company_name']
        room_name        = basic_details['room_name']
        ahu_number       = basic_details['ahu_number']
        test_taken       = datetime.datetime.today().strftime('%d/%m/%Y')
        locatiom         = basic_details['location']
        compresed_value  = basic_details['compresed_value']
        check_val        = basic_details['check_val']
        test_taken       = basic_details['Test_taken']
        regent_used      = basic_details['regent_used']
        company_name_val = company_name
        SR_NO_val        = sr_no
        test_date        = test_taken
        temp_df          = dbo.get_company_details_by_company_name(company_name_val)
        report_number    = temp_df.REPORT_NUMBER.values[0]
        customer_address = temp_df.ADDRESS.values[0]
        temp_df          = dbo.get_equipment(SR_NO_val) 
        INSTRUMENT_NAME  = temp_df.EQUIPMENT_NAME.values[0]
        MAKE             = temp_df.MAKE.values[0]
        MAKE_MODEL       = temp_df.MODEL_NUMBER.values[0]
        done_date        = temp_df.DONE_DATE.values[0].split()[0]
        due_date         = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY         = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        Nature_of_test   = "PAO REPORT"
        location         = locatiom
        customer_name    = company_name_val
        Test_taken       = test_date
        SRNO             = SR_NO_val

        compan_name = company_name_val
        compan_name = compan_name.replace(".", "")
        compan_name = compan_name.replace("/", "")
        compan_name = compan_name.replace(" ", "")

        report_counter  = dbo.get_report_number(Nature_of_test,Test_taken,str(company_name_val))
        report_counter  = report_counter+1
        temp_date = Test_taken.replace("/","")        
        report_number = str("PPE{}PAO{}{}".format(report_number,temp_date,report_counter))
        
        working_directory = MYDIR + "/" "static/Report/PAO_REPORT/{}"
        final_working_directory = "static/Report/PAO_REPORT/{}/{}.xlsx"
        file_name           = str(report_number)+"xlsx"
        #file_name = "{}_PAO_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));

        store_location = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR + "/"+final_working_directory.format(compan_name, file_name)
        print(final_working_directory)

        wb = load_workbook(os.path.join("static/inputData/Template/","PAO_template_{test_type}.xlsx".format(test_type=test_type)))
        ws = wb.active
        ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        #ws['F5'] = str(Nature_of_test)
        ws['F5'] = str(Test_taken)
        ws['F6'] = str(regent_used)
        ws['F7'] = str(ahu_number)
        ws['F8'] = str(location)
        ws['F9'] = report_number

        ws['B15'] = str(INSTRUMENT_NAME)
        ws['E15'] = str(MAKE_MODEL)
        ws['I15'] = str(SRNO)
        ws['M15'] = str(VALIDITY.replace("-", "/"))

        row = 21

        for row_data in data.itertuples():
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            ws["D" + str(row)] = row_data.INLET_NUMBER
            currentCell = ws["D" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
            ws["F" + str(row)] = row_data.Upstream
            currentCell = ws["F" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
            ws["I" + str(row)] = row_data.Leakage
            currentCell = ws["I" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
            ws["L" + str(row)] = row_data.Remark
            currentCell = ws["L" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        ws.merge_cells(start_row=21, start_column=2, end_row=row - 1, end_column=3)
        ws['B21'] = room_name

        currentCell = ws['B21']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        ws["B" + str(row)] = "Compressed Air Pressure {} kg/cm\u00B2".format(compresed_value)
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        ws["B" + str(row)] = "Acceptance criteria :Leakage should not be more Than {} %".format(str(check_val))

        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)

        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 2)] = "REPORT PREPARED BY"

        currentCell = ws["B" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 2)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)
        
        
        
        
        testtime = datetime.datetime.now(tz=gettz('Asia/Kolkata'))
        testtime = str(testtime.strftime('%d/%m/%Y %H:%M:%S')).split(" ")[0]
        ws["B" + str(row + 3)]= "Electronically Signed by {} on {} ".format(userid, testtime)
        currentCell = ws["B" + str(row + 3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEERS"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)
        testtime = datetime.datetime.now(tz=gettz('Asia/Kolkata'))
        testtime = str(testtime.strftime('%d/%m/%Y %H:%M:%S')).split(".")[0]
        #ws["B" + str(row + 9)] = "Report Prepared by {} ".format(user)


        set_border(ws, 'B1:P' + str(row + 8))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)
        dbo.insert_file(Nature_of_test,report_number,userid,Test_taken,testtime,str(company_name_val),location,final_working_directory)

        return file_name, store_location
        
    @staticmethod
    def generate_report_particle_count(data, basic_details,user,userid,EUGMP_guidlines,ISO_guidlines_master):
        sr_no = basic_details['sr_no']
        company_name = basic_details['company_name']
        room_name = basic_details['room_name']
        ahu_number = basic_details['ahu_number']
        test_taken = basic_details['Test_taken']
        locatiom = basic_details['location']
        condition = basic_details['condition']
        grade = basic_details['grade']
        gl_value = basic_details['gl_value']
        company_name_val = company_name
        SR_NO_val = sr_no
        test_date = test_taken

        temp_df = dbo.get_company_details_by_company_name(company_name_val)
        report_number = temp_df.REPORT_NUMBER.values[0]
        customer_address = temp_df.ADDRESS.values[0]
        temp_df = dbo.get_equipment(SR_NO_val) 
        INSTRUMENT_NAME = temp_df.EQUIPMENT_NAME.values[0]
        MAKE = temp_df.MAKE.values[0]
        MAKE_MODEL = temp_df.MODEL_NUMBER.values[0]
        done_date = temp_df.DONE_DATE.values[0].split()[0]
        due_date = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        Nature_of_test = "Particle Count Report"
        location = locatiom
        customer_name = company_name_val
        Test_taken = test_date
        SRNO = SR_NO_val


        compan_name = company_name_val
        compan_name = compan_name.replace(".", "")
        compan_name = compan_name.replace("/", "")
        compan_name = compan_name.replace(" ", "")
        print(gl_value)
        print(ISO_guidlines_master)
        print(EUGMP_guidlines)
        if "ISO" in gl_value :    
            trn    = ISO_guidlines_master[(ISO_guidlines_master['Grade']== grade)  ]    
        if "EU" in gl_value  :
            trn = EUGMP_guidlines[(EUGMP_guidlines['Condition'] == condition) & (EUGMP_guidlines['Grade'] == grade)]
        print(trn)
        report_counter  = dbo.get_report_number(Nature_of_test,Test_taken,str(company_name_val))
        report_counter  = report_counter+1
        temp_date = Test_taken.replace("/","")        
        report_number = str("PPE{}PC{}{}".format(report_number,temp_date,report_counter))
        
        
        
        working_directory = MYDIR + "/"  "static/Report/PARTICLE_REPORT/{}"
        final_working_directory = "static/Report/PARTICLE_REPORT/{}/{}.xlsx"
        file_name           = str(report_number)+"xlsx"
        #file_name = "{}_PARTICLE_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));
        store_location = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR +"/" + final_working_directory.format(compan_name, file_name)
        print(final_working_directory)
        if grade=="ISO 5" or grade=="A":
            wb = load_workbook(os.path.join("static/inputData/Template/",'particle_count_template_iso5.xlsx'))
        else:
            wb = load_workbook(os.path.join("static/inputData/Template/",'particle_count_template.xlsx'))
        ws = wb.active
        ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        #ws['F5'] = str(Nature_of_test)
        ws['F5'] = str(Test_taken)
        ws['F6'] = str(condition)
        if grade=="ISO 5" or grade=="A":
            ws['F7'] = str(room_name)
        else:
            ws['F7'] = str(ahu_number)
        ws['F8'] = str(location)
        #ws['F9'] = str(done_date).replace("-", "/")
        #ws['F10'] = str(due_date).replace("-", "/")
        
        ws['F9'] = report_number
        

        ws['B15'] = str(INSTRUMENT_NAME)
        ws['E15'] = str(MAKE_MODEL)
        ws['I15'] = str(SRNO)
        ws['M15'] = str(VALIDITY.replace("-", "/"))

        row = 20

        for row_data in data.itertuples():
            ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
            ws["E" + str(row)] = row_data.Location
            currentCell = ws["E" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
            ws["H" + str(row)] = row_data.zeor_point_five
            currentCell = ws["H" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
            ws["J" + str(row)] = row_data.five_point_zero
            currentCell = ws["J" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
            ws["L" + str(row)] = row_data.remark
            currentCell = ws["L" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        ws.merge_cells(start_row=20, start_column=2, end_row=row - 1, end_column=4)
        if grade=="ISO 5" or grade=="A":
            ws['B20'] = ahu_number
        else:
            ws['B20'] = room_name
        currentCell = ws['B20']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        currentCell = ws['D20']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        data.zeor_point_five = data.zeor_point_five.astype("float")
        data.five_point_zero = data.five_point_zero.astype("float")

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        ws["E" + str(row)] = "Average"
        currentCell = ws["F" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws["H" + str(row)] = round(data.zeor_point_five.mean(),0)
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        ws["J" + str(row)] = round(data.five_point_zero.mean(),0)
        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        ws["E" + str(row)] = "STD"
        currentCell = ws["F" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        if row==22:
            ws["H" + str(row)]="NA"
        else:
            ws["H" + str(row)] = round(data.zeor_point_five.std(),0)
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        
        if row==22:
            ws["J" + str(row)]="NA"
        else:    
            ws["J" + str(row)] = round(data.five_point_zero.std(),0)
        
            
        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=16)
        ws["B" + str(row)] = """Maximum Permitted number of particles/m\u00B3 equal to or greater than the tabulated size as per {GUIDELINE} """.format(GUIDELINE=gl_value)
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        row = row + 2
        
        

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        if "ISO" in gl_value : 
            ws["B" + str(row)] = "CLASS"
        else:
            ws["B" + str(row)] = "GRADE"
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
        ws["H" + str(row)] = " 0.5 \u03BC m"
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        ws["L" + str(row)] = " 5.0 \u03BCm"
        currentCell = ws["L" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        if "ISO" in gl_value : 
            ws["B" + str(row)] = "CLASS {}".format(grade.upper())
        else:
            ws["B" + str(row)] = "GRADE {}".format(grade.upper())
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
        ws["H" + str(row)] = trn['point_five_percent'].values[0]
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        ws["L" + str(row)] = trn['five_percent'].values[0]
        currentCell = ws["L" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)
        print(row)
        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 2)] = "TEST PREPARED OUT BY"

        currentCell = ws["B" + str(row + 1)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 2)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 1)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)
        
        testtime = datetime.datetime.now(tz=gettz('Asia/Kolkata'))
        testtime = str(testtime.strftime('%d/%m/%Y %H:%M:%S')).split(" ")[0]
        ws["B" + str(row + 3)]= "Electronically Signed by {} on {} ".format(userid, testtime)
        currentCell = ws["B" + str(row + 3)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEERS"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)
        testtime = datetime.datetime.now(tz=gettz('Asia/Kolkata'))
        testtime = str(testtime.strftime('%d/%m/%Y %H:%M:%S')).split(".")[0]

        #ws["B" + str(row + 9)] = "Test Prepared out by {} on {}".format(user,testtime)
        #ws["B" + str(row + 9)] = "Report Prepared by {} ".format(user)

        set_border(ws, 'B1:P' + str(row + 8))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)
        dbo.insert_file(Nature_of_test,report_number,userid,Test_taken,testtime,str(company_name_val),location,final_working_directory)
 
        return file_name, store_location
        
        
        
    @staticmethod
    def generate_thermal_report(basic_details):
    
        stages                    = ["Cycle start","Sterlization start","sterlization end","cycle End"]
        cycle_name                = str(basic_details['cycle_name'])
        Test_started_on           = str(basic_details['started_on'])
        cycle_start_time_duration = int(basic_details['cycle_start_time_duration'])
        sterlization_duration     = int(basic_details['cycle_sterlization_duration'])
        cycle_end_duration        = int(basic_details['cycle_end_duration'])
        interval_time_in_second   = int(basic_details['interval_in_seconds'])
        cycle_start_range         = [float(basic_details['cycle_start_min']),float(basic_details['cycle_start_max'])]
        cycle_end_range           = [float(basic_details['cycle_end_min']),float(basic_details['cycle_end_max'])]
        number_of_sensor          = int(basic_details['number_of_sensor'])
        format_date               = datetime.datetime.strptime(Test_started_on, '%d-%m-%Y %H:%M:%S')
        Test_conducted_on         = str(format_date).split()[0]
        Test_conducted_on         = datetime.datetime.strptime(Test_conducted_on, "%Y-%m-%d").strftime("%d-%m-%Y")
        Test_conducted_on         = str(Test_conducted_on)
        sterlization_min_range    = str(basic_details['sterlization_min']).split(",")
        sterlization_max_range    = str(basic_details['sterlization_max']).split(",")
        
       
        
        cycle_start_time = str(format_date)
        cycle_end_time1  = format_date + datetime.timedelta(minutes=cycle_start_time_duration) 
        cycle_end_time1  = str(cycle_end_time1)

        sterlization_start_time = format_date + datetime.timedelta(minutes=cycle_start_time_duration) + datetime.timedelta(seconds=interval_time_in_second)
        sterlization_start_time = str(sterlization_start_time)

        sterlization_end_time = format_date +datetime.timedelta(minutes=cycle_start_time_duration+sterlization_duration) + datetime.timedelta(seconds=interval_time_in_second)                     
        sterlization_end_time = str(sterlization_end_time)


        cycle_end_strt_time = format_date +datetime.timedelta(minutes=cycle_start_time_duration+sterlization_duration) + datetime.timedelta(seconds=interval_time_in_second)       + datetime.timedelta(seconds=interval_time_in_second)                     
        cycle_end_strt_time = str(cycle_end_strt_time)


        cycle_end_time = format_date +datetime.timedelta(minutes=cycle_start_time_duration+sterlization_duration+cycle_end_duration)                           
        cycle_end_time = str(cycle_end_time)
        cycle_start_stage = pd.date_range(cycle_start_time        , cycle_end_time1,freq="{}s".format(interval_time_in_second)).strftime('%d-%m-%Y %H:%M:%S')
        sterliztion_stage = pd.date_range(sterlization_start_time , sterlization_end_time,freq="{}s".format(interval_time_in_second)).strftime('%d-%m-%Y %H:%M:%S')
        cycle_end_stage   = pd.date_range(cycle_end_strt_time   , cycle_end_time,freq="{}s".format(interval_time_in_second)).strftime('%d-%m-%Y %H:%M:%S')
        random_value_list = [0.1,0.2,0.3,-0.1,-0.2,-0.3]
        
        record_list =[]

        range_length = len(cycle_start_stage)
        cycle_start_temp_range = np.linspace(cycle_start_range[0], cycle_start_range[1], num=range_length)
        counter = 0
        for i in cycle_start_stage:
            record = [Test_conducted_on,i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(cycle_start_temp_range[counter],cycle_start_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Cycle Started")
            record_list.append(record)
            counter=counter+1
            
            
        print(len(sterliztion_stage))
        for i in sterliztion_stage:
            record = [Test_conducted_on,i]
            counter = 0
            for x in range(number_of_sensor):              
                temp_temperture = float(random.choice(list(np.arange(float(sterlization_min_range[x]), float(sterlization_max_range[x]), 0.1))))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("In Sterlization Stage")
            record_list.append(record)
            
            counter=counter+1
    
        range_length = len(cycle_end_stage)
        cycle_end_temp_range = np.linspace(cycle_end_range[0], cycle_end_range[1], num=range_length)
        counter = 0
        for i in cycle_end_stage:
            record = [Test_conducted_on,i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(cycle_end_temp_range[counter],cycle_end_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Cycle Ended")
            record_list.append(record)
            counter=counter+1
            
         
     

        working_directory = MYDIR + "/" "static/Report/THERMAL_REPORT/{}"
        final_working_directory = "static/Report/THERMAL_REPORT/thermal.xlsx"
        file_name = "{}.xlsx".format(cycle_name)
        writer = pd.ExcelWriter(final_working_directory, engine='xlsxwriter')
        
        

        store_location = final_working_directory
        final_working_directory = MYDIR + "/"+final_working_directory
        print(final_working_directory)
        
        temp_df = pd.DataFrame(record_list)
        columns_list = ["DATE","TIME"]
        for i in  range(1,number_of_sensor+1):
            columns_list.append("CH{}".format(str(i).zfill(2)))
        columns_list.append("STAGE")
        temp_df.columns = columns_list
        for i in  range(1,number_of_sensor+1):
            columns_name = ("CH{}".format(str(i).zfill(2)))
            temp_df[columns_name] =temp_df[columns_name].astype(float)
        
        temp_df["DATE"] = "" 
        temp_df['TIME'] = temp_df['TIME'].astype(str) 
        temp_df[['DATE','TIME']] = temp_df['TIME'].str.split(' ',expand=True)
       
        #temp_df = temp_df.drop_duplicates(['DATE','TIME'],keep='last') 
        temp_df.to_excel(writer, sheet_name='Raw Data',index=False)
        trn_df = temp_df[temp_df['STAGE']=="In Sterlization Stage"]
        trn_df = trn_df.drop('STAGE', axis=1)
        
        trn_df_2 = pd.DataFrame()
        trn_df_2['MIN'] = trn_df.drop(['DATE','TIME'], axis=1).min(axis=1)
        trn_df_2['MAX'] = trn_df.drop(['DATE','TIME'], axis=1).max(axis=1)
        trn_df_2['AVG'] = trn_df.drop(['DATE','TIME'], axis=1).mean(axis=1)
        trn_df_2['AVG'] = trn_df_2['AVG'].round(1)
        
        column_list = trn_df.columns.tolist()
        action_list = ['min','max','mean']
        record_list = []
        for action in action_list:
            record = []
            record.append(action)
            for i in range(2,len(column_list)):
                operation = "round(trn_df[column_list[i]].{}(),1)".format(action)
                record.append(eval(operation))
            record_list.append(record) 
        row_counter = trn_df.shape[0]+1
        trn_df_3 = pd.DataFrame(record_list)
        
        trn_df_4 = trn_df.copy()
        delta = 60/interval_time_in_second
        for i in range(2,(len(column_list))):
            trn_df_4[column_list[i]] =trn_df_4[column_list[i]].astype(float)
            trn_df_4[column_list[i]] =  ((trn_df_4[column_list[i]]-121)/10)/delta
            trn_df_4[column_list[i]] =  10** trn_df_4[column_list[i]]
            trn_df_4[column_list[i]] =  trn_df_4[column_list[i]].round(2)
            
        
        trn_df.to_excel(writer, sheet_name='Hold_Data',index=False)
        trn_df_2.to_excel(writer, sheet_name='Hold_Data',startcol=len(column_list)+1,index=False)
        trn_df_3.to_excel(writer, sheet_name='Hold_Data',startrow=row_counter,startcol=1,header=False,index=False)
        trn_df_4.to_excel(writer, sheet_name='Hold_Data',startrow=row_counter+5,startcol=0,index=False)
        # column_list = trn_df_4.columns.tolist()
        # for i in range(2,(len(column_list))):
            # plt.plot(trn_df_4['TIME'], trn_df_4[column_list[i]])
        # plt.xticks(rotation=90)
        # plt.savefig("static/Report/THERMAL_REPORT/thermal.png",dpi=500)
        # workbook  = writer.book
        # worksheet = writer.sheets['Hold_Data']
        # worksheet.insert_image('I{}'.format(row_counter+5), "static/Report/THERMAL_REPORT/thermal.png")
        
        writer.close()

        return file_name, store_location

    @staticmethod
    def generate_thermal_report_new(basic_details):
        random_value_list = [0.1,0.2,0.3,-0.1,-0.2,-0.3]
        tunnel_temp_range = 98.9
        record_list = []
        
        report_name               = str(basic_details['report_name'])
        number_of_sensor          = int(basic_details['number_of_sensor'])
        interval_in_seconds       = int(basic_details['interval_in_seconds'])
        
        stage1_start_time         = str(basic_details['stage1_start_time'])
        stage1_end_time           = str(basic_details['stage1_end_time'])
        stage1_start_temperature  = float(basic_details['stage1_start_temperature'])
        stage1_end_temperature    = float(basic_details['stage1_end_temperature'])
        
        
        
        number_of_left_side_pulse   = int(basic_details['number_of_left_side_pulse'])
        left_side_pulse_start_time  = str(basic_details['left_side_pulse_start_time'])
        left_side_pulse_end_time    = str(basic_details['left_side_pulse_end_time'])
        left_side_pulse_start_temperature = float(basic_details['left_side_pulse_start_temperature'])
        left_side_pulse_end_temperature   = float(basic_details['left_side_pulse_end_temperature'])
        
        
        
        
        pre_sterlization_start_time  = str(basic_details['pre_sterlization_start_time'])
        pre_sterlization_end_time    = str(basic_details['pre_sterlization_end_time'])
        pre_sterlization_start_temperature = float(basic_details['pre_sterlization_start_temperature'])
        pre_sterlization_end_temperature   = float(basic_details['pre_sterlization_end_temperature'])
        
        sterlization_start_time  = str(basic_details['sterlization_start_time'])
        sterlization_end_time    = str(basic_details['sterlization_end_time'])
        sterlization_start_temperature = float(basic_details['sterlization_start_temperature'])
        sterlization_end_temperature   = float(basic_details['sterlization_end_temperature'])
        
        post_sterlization_start_time  = str(basic_details['post_sterlization_start_time'])
        post_sterlization_end_time    = str(basic_details['post_sterlization_end_time'])
        post_sterlization_start_temperature = float(basic_details['post_sterlization_start_temperature'])
        post_sterlization_end_temperature   = float(basic_details['post_sterlization_end_temperature'])
        
        
        
        
        number_of_right_side_pulse   = int(basic_details['number_of_right_side_pulse'])
        right_side_pulse_start_time  = str(basic_details['right_side_pulse_start_time'])
        right_side_pulse_end_time    = str(basic_details['right_side_pulse_end_time'])
        right_side_pulse_start_temperature = float(basic_details['right_side_pulse_start_temperature'])
        right_side_pulse_end_temperature   = float(basic_details['right_side_pulse_end_temperature'])
        
        
        
        
        final_stage_start_time         = str(basic_details['final_stage_start_time'])
        final_stage_end_time           = str(basic_details['final_stage_end_time'])
        final_stage_start_temperature  = float(basic_details['final_stage_start_temperature'])
        final_stage_end_temperature    = float(basic_details['final_stage_end_temperature'])

        cycle_start_stage = pd.date_range(stage1_start_time,stage1_end_time,freq="{}s".format(interval_in_seconds)).strftime('%d-%m-%Y %H:%M:%S')
        range_length      = len(cycle_start_stage)
        cycle_start_temp_range = np.linspace(stage1_start_temperature, stage1_end_temperature, num=range_length)
        counter = 0
        for i in cycle_start_stage:
            record = ["",i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(cycle_start_temp_range[counter],cycle_start_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Cycle Started")
            record_list.append(record)
            counter=counter+1

        if number_of_left_side_pulse>0:
            left_side_peak_stage = pd.date_range(left_side_pulse_start_time,left_side_pulse_end_time,freq="{}s".format(interval_in_seconds)).strftime('%d-%m-%Y %H:%M:%S')
            
            for temp_peak_range in np.array_split(left_side_peak_stage, number_of_left_side_pulse):
                low_to_high_range , high_to_low_range = np.array_split(temp_peak_range, 2)
            
                range_length      = len(low_to_high_range)
                temp_range = np.linspace(left_side_pulse_start_temperature, left_side_pulse_end_temperature, num=range_length)
                counter = 0
                for i in low_to_high_range:
                    record = ["",i]
                    if counter  == range_length-1:
                        counter = counter-1
                    for x in range(number_of_sensor):
                        temp_temperture = random.uniform(temp_range[counter],temp_range[counter+1])
                        temp_temperture = temp_temperture + float(random.choice(random_value_list))
                        temp_temperture = '{0:.1f}'.format(temp_temperture) 
                        record.append(temp_temperture)
                    record.append("PEAK CYCLE")
                    record_list.append(record)
                    counter=counter+1
            

                range_length      = len(high_to_low_range)
                temp_range = np.linspace(left_side_pulse_end_temperature, left_side_pulse_start_temperature, num=range_length)
                counter = 0
                for i in high_to_low_range:
                    record = ["",i]
                    if counter  == range_length-1:
                        counter = counter-1
                    for x in range(number_of_sensor):
                        temp_temperture = random.uniform(temp_range[counter],temp_range[counter+1])
                        temp_temperture = temp_temperture + float(random.choice(random_value_list))
                        temp_temperture = '{0:.1f}'.format(temp_temperture) 
                        record.append(temp_temperture)
                    record.append("PEAK CYCLE")
                    record_list.append(record)
                    counter=counter+1

        sterlization_stage      = pd.date_range(pre_sterlization_start_time,pre_sterlization_end_time,freq="{}s".format(interval_in_seconds)).strftime('%d-%m-%Y %H:%M:%S')
        range_length            = len(sterlization_stage)
        sterlization_temp_range = np.linspace(pre_sterlization_start_temperature, pre_sterlization_end_temperature, num=range_length)
        counter = 0
        for i in sterlization_stage:
            record = ["",i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(sterlization_temp_range[counter],sterlization_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Pre Sterlization Stage")
            record_list.append(record)
            counter=counter+1
        
        
        
        
        
        sterlization_stage      = pd.date_range(sterlization_start_time,sterlization_end_time,freq="{}s".format(interval_in_seconds)).strftime('%d-%m-%Y %H:%M:%S')
        range_length            = len(sterlization_stage)
        sterlization_temp_range = np.linspace(sterlization_start_temperature, sterlization_end_temperature, num=range_length)
        counter = 0
        for i in sterlization_stage:
            record = ["",i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(sterlization_temp_range[counter],sterlization_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Sterlization Stage")
            record_list.append(record)
            counter=counter+1
        
        
        sterlization_stage      = pd.date_range(post_sterlization_start_time,post_sterlization_end_time,freq="{}s".format(interval_in_seconds)).strftime('%d-%m-%Y %H:%M:%S')
        range_length            = len(sterlization_stage)
        sterlization_temp_range = np.linspace(post_sterlization_start_temperature, post_sterlization_end_temperature, num=range_length)
        counter = 0
        for i in sterlization_stage:
            record = ["",i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(sterlization_temp_range[counter],sterlization_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Post Sterlization Stage")
            record_list.append(record)
            counter=counter+1
        
        
        if number_of_right_side_pulse >0:
            right_side_peak_stage = pd.date_range(right_side_pulse_start_time,right_side_pulse_end_time,freq="{}s".format(interval_in_seconds)).strftime('%d-%m-%Y %H:%M:%S')
            
            for temp_peak_range in np.array_split(right_side_peak_stage, number_of_right_side_pulse):
                low_to_high_range , high_to_low_range = np.array_split(temp_peak_range, 2)
            
                range_length      = len(low_to_high_range)
                temp_range = np.linspace(right_side_pulse_start_temperature, right_side_pulse_end_temperature, num=range_length)
                counter = 0
                for i in low_to_high_range:
                    record = ["",i]
                    if counter  == range_length-1:
                        counter = counter-1
                    for x in range(number_of_sensor):
                        temp_temperture = random.uniform(temp_range[counter],temp_range[counter+1])
                        temp_temperture = temp_temperture + float(random.choice(random_value_list))
                        temp_temperture = '{0:.1f}'.format(temp_temperture) 
                        record.append(temp_temperture)
                    record.append("PEAK CYCLE")
                    record_list.append(record)
                    counter=counter+1
            
                range_length      = len(high_to_low_range)
                temp_range = np.linspace(right_side_pulse_end_temperature, right_side_pulse_start_temperature, num=range_length)
                counter = 0
                for i in high_to_low_range:
                    record = ["",i]
                    if counter  == range_length-1:
                        counter = counter-1
                    for x in range(number_of_sensor):
                        temp_temperture = random.uniform(temp_range[counter],temp_range[counter+1])
                        temp_temperture = temp_temperture + float(random.choice(random_value_list))
                        temp_temperture = '{0:.1f}'.format(temp_temperture) 
                        record.append(temp_temperture)
                    record.append("PEAK CYCLE")
                    record_list.append(record)
                    counter=counter+1


        cycle_final_stage = pd.date_range(final_stage_start_time,final_stage_end_time,freq="{}s".format(interval_in_seconds)).strftime('%d-%m-%Y %H:%M:%S')
        range_length      = len(cycle_final_stage)
        cycle_final_temp_range = np.linspace(final_stage_start_temperature, final_stage_end_temperature, num=range_length)
        counter = 0
        for i in cycle_final_stage:
            record = ["",i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(cycle_final_temp_range[counter],cycle_final_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Cycle Ended")
            record_list.append(record)
            counter=counter+1


        temp_df = pd.DataFrame(record_list)
        columns_list = ["DATE","TIME"]
        for i in  range(1,number_of_sensor+1):
            columns_list.append("CH{}".format(str(i).zfill(2)))
        columns_list.append("STAGE")
        temp_df.columns = columns_list
        for i in  range(1,number_of_sensor+1):
            columns_name = ("CH{}".format(str(i).zfill(2)))
            temp_df[columns_name] =temp_df[columns_name].astype(float)
        
        temp_df["DATE"] = "" 
        temp_df['TIME'] = temp_df['TIME'].astype(str) 
        temp_df[['DATE','TIME']] = temp_df['TIME'].str.split(' ',expand=True)



        working_directory = MYDIR + "/" "static/Report/THERMAL_REPORT/{}"
        final_working_directory = "static/Report/THERMAL_REPORT/thermal.xlsx"
        file_name = "{}.xlsx".format(report_name)
        writer = pd.ExcelWriter(final_working_directory, engine='xlsxwriter')

        store_location = final_working_directory
        final_working_directory = MYDIR + "/"+final_working_directory
        print(final_working_directory)

        
        temp_df.to_excel(writer, sheet_name='Raw Data',index=False)
        trn_df = temp_df[temp_df['STAGE']=="Sterlization Stage"]
        trn_df = trn_df.drop('STAGE', axis=1)
        
        trn_df_2 = pd.DataFrame()
        trn_df_2['MIN'] = trn_df.drop(['DATE','TIME'], axis=1).min(axis=1)
        trn_df_2['MAX'] = trn_df.drop(['DATE','TIME'], axis=1).max(axis=1)
        trn_df_2['AVG'] = trn_df.drop(['DATE','TIME'], axis=1).mean(axis=1)
        trn_df_2['DIF'] = trn_df_2['MAX'] - trn_df_2['MIN']
        trn_df_2['AVG'] = trn_df_2['AVG'].round(1)
        
        column_list = trn_df.columns.tolist()
        action_list = ['min','max','mean','dif']
        record_list = []
        for action in action_list:
            record = []
            record.append(action)
            if action!="dif":
                for i in range(2,len(column_list)):
                    operation = "round(trn_df[column_list[i]].{}(),1)".format(action)
                    record.append(eval(operation))
            if action == "dif":
                for i in range(2,len(column_list)):
                    max = "round(trn_df[column_list[i]].max(),1)"
                    max = eval(max)
                    min = "round(trn_df[column_list[i]].min(),1)"
                    min = eval(min)
                    dif = max-min
                    record.append(dif)
                
            record_list.append(record) 
        row_counter = trn_df.shape[0]+1
        trn_df_3 = pd.DataFrame(record_list)
        
        trn_df_4 = trn_df.copy()
        delta = 60/interval_in_seconds
        for i in range(2,(len(column_list))):
            trn_df_4[column_list[i]] =trn_df_4[column_list[i]].astype(float)
            trn_df_4[column_list[i]] =  ((trn_df_4[column_list[i]]-121)/10)/delta
            trn_df_4[column_list[i]] =  pow(10,trn_df_4[column_list[i]])
            trn_df_4[column_list[i]] =  trn_df_4[column_list[i]].round(2)
            
        
        trn_df.to_excel(writer, sheet_name='Hold_Data',index=False)
        trn_df_2.to_excel(writer, sheet_name='Hold_Data',startcol=len(column_list)+1,index=False)
        trn_df_3.to_excel(writer, sheet_name='Hold_Data',startrow=row_counter,startcol=1,header=False,index=False)
        trn_df_4.to_excel(writer, sheet_name='Hold_Data',startrow=row_counter+5,startcol=0,index=False)
        
        writer.close()
        return file_name, store_location
            
            





