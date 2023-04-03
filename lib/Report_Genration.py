import pandas as pd
from flask import Flask, render_template, flash, url_for, request, make_response, jsonify, session, send_from_directory
from werkzeug.utils import secure_filename
import os, time
import io
import pytz
import base64
import json
import datetime
import xlsxwriter
import os, sys, glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles import Border, Side
# from flask import send_from_directory
from flask import send_file
import smtplib, ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from fnmatch import fnmatch
import time
from db_dervices import DBO
import pandas as pd
import numpy as np
import datetime
import random
from dateutil.tz import gettz

dbo          =  DBO()
MYDIR        = os.path.dirname(__file__)
REGULAR_SIZE = 11
REGULAR_FONT = 'Cambria'
server       = 'smtp.gmail.com'
port         = 587
username     = "aajeetshk@gmail.com"
password     = "ilbumnmnsnqletdk"
send_from    = "aajeetshk@gmail.com"
send_to      = "ashish@pinpointengineers.co.in"


def sum_velocty(row):
    sum = int(row.V1) + int(row.V2) + int(row.V3) + int(row.V4) + int(row.V5)
    avg = sum / 5
    return int(avg)


def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            



class Report_Genration:
    def __init__(self):
        self = self

    @staticmethod
    def generate_report_air_velocity(data, basic_details,user):
        sr_no = basic_details['sr_no']
        company_name = basic_details['company_name']
        room_volume = basic_details['room_volume']
        room_name = basic_details['room_name']
        ahu_number = basic_details['ahu_number']
        test_taken = basic_details['Test_taken']
        locatiom = basic_details['location']
        grade = basic_details['grade']
        acph_thresold = basic_details['acph_thresold']
        company_name_val = company_name
        SR_NO_val = sr_no
        test_date = test_taken

        temp_df = dbo.get_company_details_by_company_name(company_name)
        report_number = temp_df.REPORT_NUMBER.values[0]
        customer_address = temp_df.ADDRESS.values[0]
        temp_df = dbo.get_equipment_by_id(SR_NO_val) 
        INSTRUMENT_NAME = temp_df.EQUIPMENT_NAME.values[0]
        MAKE = temp_df.MAKE.values[0]
        MAKE_MODEL = temp_df.MODEL_NUMBER.values[0]
        done_date = temp_df.DONE_DATE.values[0].split()[0]
        due_date = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        Nature_of_test = "AIR VELOCITY"
        location = locatiom
        customer_name = company_name_val
        Test_taken = test_date
        ahu_number = ahu_number
        SRNO = SR_NO_val

        compan_name = company_name_val
        compan_name = compan_name.replace(".", "")
        compan_name = compan_name.replace("/", "")
        compan_name = compan_name.replace(" ", "")

        working_directory = MYDIR + "/" +"static/Report/AIR_VELOCITY_REPORT/{}"
        final_working_directory = "static/Report/AIR_VELOCITY_REPORT/{}/{}.xlsx"
        file_name = "{}_AIR_VELOCOTY_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));

        store_location = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR + "/" +final_working_directory.format(compan_name, file_name)
        print(final_working_directory)

 
        wb = load_workbook(os.path.join("static/inputData/Template/",'Air_velocity_template.xlsx'))
        ws = wb.active
        ws.protection.sheet = True

        if grade=="A":
            wb = load_workbook(os.path.join("static/inputData/Template/",'Air_velocity_template_grade_a.xlsx'))
            ws = wb.active
            ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        ws['F5'] = str(Nature_of_test)
        ws['F6'] = str(Test_taken)
        ws['F7'] = str(ahu_number)
        ws['F8'] = str(location)
        ws['F9'] = str("PPE0{}AV01A".format(report_number))

        ws['B15'] = str(INSTRUMENT_NAME)
        ws['E15'] = str(MAKE_MODEL)
        ws['I15'] = str(SRNO)
        ws['M15'] = str(VALIDITY.replace("-", "/"))

        ws.merge_cells(start_row=21, start_column=2, end_row=21, end_column=3)
        # ws.merge_cells(start_row=23, start_column=3, end_row=23, end_column=4)
        ws.merge_cells(start_row=21, start_column=15, end_row=21, end_column=16)

        row = 22
        data['AVG_Velocity'] = data.apply(sum_velocty, axis=1)
        if grade=="A":
            data['CFM'] = data.apply(lambda x: int(x['AVG_Velocity']), axis=1)
            Total_cfm   = data['CFM'].sum()
            ACPH_VALUE  = float((Total_cfm * 60))
            ACPH_VALUE  = int(ACPH_VALUE)
        else:
            data['CFM'] = data.apply(lambda x: int(x['AVG_Velocity']) * int(x['Inlet_size']), axis=1)
            Total_cfm   = data['CFM'].sum()
            ACPH_VALUE  = float((Total_cfm * 60)) / float(room_volume)
            ACPH_VALUE  = int(ACPH_VALUE)

        # set_border(ws, "B16:M17")
        for row_data in data.itertuples():
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            ws["D" + str(row)] = row_data.Label_number
            currentCell = ws["D" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')
          
            if grade=="A":
                ws["E" + str(row)] = row_data.V1
                ws["G" + str(row)] = row_data.V2
                ws["I" + str(row)] = row_data.V3
                ws["K" + str(row)] = row_data.V4
                ws["M" + str(row)] = row_data.V5
                ws["O" + str(row)] = row_data.AVG_Velocity

            else:
                ws["F" + str(row)] = row_data.V1
                ws["G" + str(row)] = row_data.V2
                ws["H" + str(row)] = row_data.V3
                ws["I" + str(row)] = row_data.V4
                ws["J" + str(row)] = row_data.V5
                ws["K" + str(row)] = row_data.AVG_Velocity
                ws["L" + str(row)] = row_data.Inlet_size
                ws["M" + str(row)] = row_data.CFM
        
            row += 1

        ws.merge_cells(start_row=22, start_column=2, end_row=row - 1, end_column=3)
        ws['B22'] = room_name
        
        if grade=="A":
            pass
        else:
            ws.merge_cells(start_row=22, start_column=14, end_row=row - 1, end_column=14)
            ws['N22'] = room_volume
            ws.merge_cells(start_row=22, start_column=15, end_row=row - 1, end_column=16)
            ws['O22'] = ACPH_VALUE

        currentCell = ws['B22']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        currentCell = ws['N22']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        currentCell = ws['O22']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        if grade!="A":
           ws["B" + str(row)] = "TOTAL CFM"
           currentCell = ws["B" + str(row)]
           currentCell.alignment = Alignment(horizontal='right', vertical='center')

           ws.merge_cells(start_row=row, start_column=13, end_row=row, end_column=13)
           ws["M" + str(row)] = Total_cfm
           ws.merge_cells(start_row=row, start_column=15, end_row=row, end_column=16)

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        ws["B" + str(row)] = "Acceptance criteria :Not less Than {} ACPH".format(str(acph_thresold))

        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=16)
        ws["J" + str(row)] = "GRADE : {}".format(grade)

        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)

        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 2)] = "TEST CARRIED OUT BY"

        currentCell = ws["B" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 2)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEERS"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)
        
       
        testtime = datetime.datetime.now(tz=gettz('Asia/Kolkata'))
        testtime = str(testtime.strftime('%d/%m/%Y %H:%M:%S')).split(".")[0]

        ws["B" + str(row + 9)] = "Test Carried out by {} on {}".format(user, testtime)

        set_border(ws, 'B1:P' + str(row + 9))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)
        #dbo.insert_file(file_name,final_working_directory)
        return file_name, store_location

    @staticmethod
    def generate_report_pao(data, basic_details,user):
        sr_no = basic_details['sr_no']
        company_name = basic_details['company_name']
        room_name = basic_details['room_name']
        ahu_number = basic_details['ahu_number']

        test_taken = basic_details['Test_taken']
        locatiom = basic_details['location']
        compresed_value = basic_details['compresed_value']
        check_val = basic_details['check_val']
        company_name_val = company_name
        SR_NO_val = sr_no
        test_date = test_taken

        temp_df = dbo.get_company_details_by_company_name(company_name_val)
        report_number = temp_df.REPORT_NUMBER.values[0]
        customer_address = temp_df.ADDRESS.values[0]
        temp_df = dbo.get_equipment_by_id(SR_NO_val) 
        INSTRUMENT_NAME = temp_df.EQUIPMENT_NAME.values[0]
        MAKE = temp_df.MAKE.values[0]
        MAKE_MODEL = temp_df.MODEL_NUMBER.values[0]
        done_date = temp_df.DONE_DATE.values[0].split()[0]
        due_date = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        Nature_of_test = "PAO REPORT"
        location = locatiom
        customer_name = company_name_val
        Test_taken = test_date
        SRNO = SR_NO_val

        compan_name = company_name_val
        compan_name = compan_name.replace(".", "")
        compan_name = compan_name.replace("/", "")
        compan_name = compan_name.replace(" ", "")

        working_directory = MYDIR + "/" "static/Report/PAO_REPORT/{}"
        final_working_directory = "static/Report/PAO_REPORT/{}/{}.xlsx"
        file_name = "{}_PAO_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));

        store_location = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR + "/"+final_working_directory.format(compan_name, file_name)
        print(final_working_directory)

        wb = load_workbook(os.path.join("static/inputData/Template/",'PAO_template.xlsx'))
        ws = wb.active
        ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        ws['F5'] = str(Nature_of_test)
        ws['F6'] = str(Test_taken)
        ws['F7'] = str(ahu_number)
        ws['F8'] = str(location)
        ws['F9'] = str(done_date).replace("-", "/")
        ws['F10'] = str(due_date).replace("-", "/")
        ws['F11'] = str("PPE0{}AV01A".format(report_number))

        ws['B17'] = str(INSTRUMENT_NAME)
        ws['E17'] = str(MAKE_MODEL)
        ws['I17'] = str(SRNO)
        ws['M17'] = str(VALIDITY.replace("-", "/"))

        row = 24

        for row_data in data.itertuples():
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            ws["D" + str(row)] = row_data.INLET_NUMBER
            currentCell = ws["D" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
            ws["F" + str(row)] = row_data.Upstream
            currentCell = ws["F" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=11)
            ws["I" + str(row)] = row_data.Leakage
            currentCell = ws["I" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
            ws["L" + str(row)] = row_data.Remark
            currentCell = ws["L" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        ws.merge_cells(start_row=24, start_column=2, end_row=row - 1, end_column=3)
        ws['B24'] = room_name

        currentCell = ws['B24']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        ws["B" + str(row)] = "Compressed Air Presure {}".format(compresed_value)
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)
        ws["B" + str(row)] = "Acceptance criteria :Not less Than {} %".format(str(check_val))

        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)

        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 2)] = "TEST CARRIED OUT BY"

        currentCell = ws["B" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 2)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 2)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEER"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)
        ws["B" + str(row + 9)] = "Test Carried out by {} on {}".format(user,datetime.datetime.today().strftime('%d/%m/%Y %H:%M:%S'))


        set_border(ws, 'B1:P' + str(row + 9))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)

        return file_name, store_location

    @staticmethod
    def generate_report_particle_count(data, basic_details,user,EUGMP_guidlines,ISO_guidlines_master):
        sr_no = basic_details['sr_no']
        company_name = basic_details['company_name']
        room_name = basic_details['room_name']
        ahu_number = basic_details['ahu_number']
        test_taken = basic_details['Test_taken']
        locatiom = basic_details['location']
        condition = basic_details['condition']
        grade = basic_details['grade']
        gl_value = basic_details['gl_value']
        company_name_val = company_name
        SR_NO_val = sr_no
        test_date = test_taken

        temp_df = dbo.get_company_details_by_company_name(company_name_val)
        report_number = temp_df.REPORT_NUMBER.values[0]
        customer_address = temp_df.ADDRESS.values[0]
        temp_df = dbo.get_equipment_by_id(SR_NO_val) 
        INSTRUMENT_NAME = temp_df.EQUIPMENT_NAME.values[0]
        MAKE = temp_df.MAKE.values[0]
        MAKE_MODEL = temp_df.MODEL_NUMBER.values[0]
        done_date = temp_df.DONE_DATE.values[0].split()[0]
        due_date = temp_df.DUE_DATE.values[0].split()[0]
        VALIDITY = temp_df.DUE_DATE.values[0].replace("-","/").split()[0]
        Nature_of_test = "PARTICLE COUNT REPORT"
        location = locatiom
        customer_name = company_name_val
        Test_taken = test_date
        SRNO = SR_NO_val


        compan_name = company_name_val
        compan_name = compan_name.replace(".", "")
        compan_name = compan_name.replace("/", "")
        compan_name = compan_name.replace(" ", "")
        print(gl_value)
        print(ISO_guidlines_master)
        print(EUGMP_guidlines)
        if "ISO" in gl_value :    
            trn    = ISO_guidlines_master[(ISO_guidlines_master['Grade']== grade)  ]    
        if "EU" in gl_value  :
            trn = EUGMP_guidlines[(EUGMP_guidlines['Condition'] == condition) & (EUGMP_guidlines['Grade'] == grade)]
        print(trn)
        working_directory = MYDIR + "/"  "static/Report/PARTICLE_REPORT/{}"
        final_working_directory = "static/Report/PARTICLE_REPORT/{}/{}.xlsx"
        file_name = "{}_PARTICLE_REPORT_{}".format(room_name, str(datetime.datetime.today().strftime('%d_%m_%Y')))
        if not os.path.exists(working_directory.format(compan_name)):
            os.mkdir(working_directory.format(compan_name));
        store_location = final_working_directory.format(compan_name, file_name)
        final_working_directory = MYDIR +"/" + final_working_directory.format(compan_name, file_name)
        print(final_working_directory)

        wb = load_workbook(os.path.join("static/inputData/Template/",'particle_count_template.xlsx'))
        ws = wb.active
        #ws.protection.sheet = True

        # Data can be assigned directly to cells
        ws['F3'] = str(company_name_val)
        ws['F4'] = customer_address
        ws['F5'] = str(Nature_of_test)
        ws['F6'] = str(Test_taken)
        ws['F7'] = str(ahu_number)
        ws['F8'] = str(location)
        #ws['F9'] = str(done_date).replace("-", "/")
        #ws['F10'] = str(due_date).replace("-", "/")
        ws['F9'] = str("PPE0{}AV01A".format(report_number))

        ws['B15'] = str(INSTRUMENT_NAME)
        ws['E15'] = str(MAKE_MODEL)
        ws['I15'] = str(SRNO)
        ws['M15'] = str(VALIDITY.replace("-", "/"))

        row = 21

        for row_data in data.itertuples():
            ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
            ws["F" + str(row)] = row_data.Location
            currentCell = ws["F" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
            ws["H" + str(row)] = row_data.zeor_point_five
            currentCell = ws["H" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
            ws["J" + str(row)] = row_data.five_point_zero
            currentCell = ws["J" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
            ws["L" + str(row)] = row_data.remark
            currentCell = ws["L" + str(row)]
            currentCell.alignment = Alignment(horizontal='center', vertical='center')

            row += 1

        ws.merge_cells(start_row=21, start_column=4, end_row=row - 1, end_column=5)
        ws['D21'] = room_name
        ws.merge_cells(start_row=21, start_column=2, end_row=row - 1, end_column=3)
        ws['B23'] = "1"
        currentCell = ws['B21']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        currentCell = ws['D21']
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        data.zeor_point_five = data.zeor_point_five.astype("float")
        data.five_point_zero = data.five_point_zero.astype("float")

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws["F" + str(row)] = "Average"
        currentCell = ws["F" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws["H" + str(row)] = data.zeor_point_five.mean()
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        ws["J" + str(row)] = data.five_point_zero.mean()
        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=7)
        ws["F" + str(row)] = "STD"
        currentCell = ws["F" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
        ws["H" + str(row)] = data.zeor_point_five.std()
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
        ws["J" + str(row)] = data.five_point_zero.std()
        currentCell = ws["J" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=16)
        ws["B" + str(row)] = "Maximum {}".format(gl_value)
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')
        row = row + 2

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws["B" + str(row)] = "GRADE"
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
        ws["H" + str(row)] = " 0.5 m"
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        ws["L" + str(row)] = " 5.0 m"
        currentCell = ws["L" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1

        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
        ws["B" + str(row)] = "GRADE {}".format(grade.upper())
        currentCell = ws["B" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=11)
        ws["H" + str(row)] = trn['point_five_percent'].values[0]
        currentCell = ws["H" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells(start_row=row, start_column=12, end_row=row, end_column=16)
        ws["L" + str(row)] = trn['five_percent'].values[0]
        currentCell = ws["L" + str(row)]
        currentCell.alignment = Alignment(horizontal='center', vertical='center')

        row = row + 1
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=16)

        ws.merge_cells(start_row=(row + 1), start_column=2, end_row=(row + 1), end_column=7)
        ws.merge_cells(start_row=(row + 1), start_column=8, end_row=(row + 1), end_column=16)

        ws.merge_cells(start_row=(row + 2), start_column=2, end_row=(row + 2), end_column=7)
        ws.merge_cells(start_row=(row + 2), start_column=8, end_row=(row + 2), end_column=16)

        ws["B" + str(row + 1)] = "TEST CARRIED OUT BY"

        currentCell = ws["B" + str(row + 1)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 1)] = "VERIFIED BY"

        currentCell = ws["H" + str(row + 1)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws.merge_cells(start_row=(row + 3), start_column=2, end_row=(row + 7), end_column=7)
        ws.merge_cells(start_row=(row + 3), start_column=8, end_row=(row + 7), end_column=16)

        # if data.shape[0]==1:
        #    row =row +1
        #####################################################################
        ws.merge_cells(start_row=(row + 8), start_column=2, end_row=(row + 8), end_column=7)
        ws.merge_cells(start_row=(row + 8), start_column=8, end_row=(row + 8), end_column=16)

        ws["B" + str(row + 8)] = "PIN POINT ENGINEER"

        currentCell = ws["B" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')

        ws["H" + str(row + 8)] = company_name_val

        currentCell = ws["H" + str(row + 8)]
        currentCell.alignment = Alignment(horizontal='center', vertical='top')
        #######################################
        
        ws.merge_cells(start_row=(row + 9), start_column=2, end_row=(row + 9), end_column=16)

        ws["B" + str(row + 9)] = "Test Carried out by {} on {}".format(user,datetime.datetime.today(pytz.UTC).strftime('%d/%m/%Y %H:%M:%S'))

        set_border(ws, 'B1:P' + str(row + 8))
        set_border(ws, 'B1:P' + str(1))

        wb.save(final_working_directory)

        return file_name, store_location
        
        
    @staticmethod
    def generate_thermal_report(basic_details):
    
        stages                    = ["Cycle start","Sterlization start","sterlization end","cycle End"]
        cycle_name                = str(basic_details['cycle_name'])
        Test_started_on           = str(basic_details['started_on'])
        cycle_start_time_duration = int(basic_details['cycle_start_time_duration'])
        sterlization_duration     = int(basic_details['cycle_sterlization_duration'])
        cycle_end_duration        = int(basic_details['cycle_end_duration'])
        interval_time_in_second   = int(basic_details['interval_in_seconds'])
        cycle_start_range         = [float(basic_details['cycle_start_min']),float(basic_details['cycle_start_max'])]
        sterlization_range        = [float(basic_details['sterlization_min']),float(basic_details['sterlization_max'])]
        cycle_end_range           = [float(basic_details['cycle_end_min']),float(basic_details['cycle_end_max'])]
        number_of_sensor          = int(basic_details['number_of_sensor'])
        format_date               = datetime.datetime.strptime(Test_started_on, '%d-%m-%Y %H:%M:%S')
        Test_conducted_on         = str(format_date).split()[0]
        Test_conducted_on         = datetime.datetime.strptime(Test_conducted_on, "%Y-%m-%d").strftime("%d-%m-%Y")
        Test_conducted_on         = str(Test_conducted_on)
       
        
        cycle_start_time = str(format_date).split()[1]
        sterlization_start_time = format_date + datetime.timedelta(minutes=cycle_start_time_duration)
        sterlization_start_time = str(sterlization_start_time).split()[1]

        sterlization_end_time = format_date +datetime.timedelta(minutes=cycle_start_time_duration+sterlization_duration)                           
        sterlization_end_time = str(sterlization_end_time).split()[1]


        cycle_end_time = format_date +datetime.timedelta(minutes=cycle_start_time_duration+sterlization_duration+cycle_end_duration)                           
        cycle_end_time = str(cycle_end_time).split()[1]
        cycle_start_stage = pd.date_range(cycle_start_time        , sterlization_start_time,freq="{}s".format(interval_time_in_second)).strftime('%H:%M:%S')
        sterliztion_stage = pd.date_range(sterlization_start_time , sterlization_end_time,freq="{}s".format(interval_time_in_second)).strftime('%H:%M:%S')
        cycle_end_stage   = pd.date_range(sterlization_end_time   , cycle_end_time,freq="{}s".format(interval_time_in_second)).strftime('%H:%M:%S')
        random_value_list = [0.1,0.2,0.3,-0.1,-0.2,-0.3]
        
        record_list =[]

        range_length = len(cycle_start_stage)
        cycle_start_temp_range = np.linspace(cycle_start_range[0], cycle_start_range[1], num=range_length)
        counter = 0
        for i in cycle_start_stage:
            record = [Test_conducted_on,i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(cycle_start_temp_range[counter],cycle_start_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Cycle Started")
            record_list.append(record)
            counter=counter+1

        for i in sterliztion_stage:
            record = [Test_conducted_on,i]
            counter = 0
            for x in range(number_of_sensor):
                temp_temperture = float(random.choice(list(np.arange(sterlization_range[0], sterlization_range[1], 0.1))))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("In Sterlization Stage")
            record_list.append(record)
            counter=counter+1
    
        range_length = len(cycle_end_stage)
        cycle_end_temp_range = np.linspace(cycle_end_range[0], cycle_end_range[1], num=range_length)
        counter = 0
        for i in cycle_end_stage:
            record = [Test_conducted_on,i]
            if counter  == range_length-1:
                counter = counter-1
            for x in range(number_of_sensor):
                temp_temperture = random.uniform(cycle_end_temp_range[counter],cycle_end_temp_range[counter+1])
                temp_temperture = temp_temperture + float(random.choice(random_value_list))
                temp_temperture = '{0:.1f}'.format(temp_temperture) 
                record.append(temp_temperture)
            record.append("Cycle Ended")
            record_list.append(record)
            counter=counter+1
            
         


        working_directory = MYDIR + "/" "static/Report/THERMAL_REPORT/{}"
        final_working_directory = "static/Report/THERMAL_REPORT/thermal.xlsx"
        file_name = "{}.xlsx".format(cycle_name)#.format(str(datetime.datetime.today().strftime('%d_%m_%Y')))
        

        store_location = final_working_directory
        final_working_directory = MYDIR + "/"+final_working_directory
        print(final_working_directory)
        
        temp_df = pd.DataFrame(record_list)
        columns_list = ["DATE","TIME"]
        for i in  range(1,number_of_sensor+1):
            columns_list.append("CH{}".format(str(i).zfill(2)))
        columns_list.append("STAGE")
        temp_df.columns = columns_list
        for i in  range(1,number_of_sensor+1):
            columns_name = ("CH{}".format(str(i).zfill(2)))
            temp_df[columns_name] =temp_df[columns_name].astype(float)
        
        temp_df["DATE"] = temp_df["DATE"].astype(str)
        temp_df.to_excel(final_working_directory,index=False)

        return file_name, store_location


